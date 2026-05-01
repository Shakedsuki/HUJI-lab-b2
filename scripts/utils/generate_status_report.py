"""
generate_status_report.py
--------------------------
Reads data/experiments.json and scans measurements/ to produce a single
Excel workbook (default: data/status_report.xlsx) summarising every
measurement: ICs, status, dropout statistics, file presence, and
human-readable notes.

The script is idempotent — it always regenerates the workbook from
scratch from the current state of experiments.json + measurements/.
Dropout numbers are computed directly from each tracking.csv (the CSV
is the ground truth, not the registry).

Usage
~~~~~
    python scripts/utils/generate_status_report.py
    python scripts/utils/generate_status_report.py --output data/my_report.xlsx
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

# Print summary line uses dashes / unicode in some places — force UTF-8
# so the script behaves the same in cmd, PowerShell, Git Bash, and pipes.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except (AttributeError, OSError):
    pass

# ── External deps: openpyxl + pandas. Fail with a helpful message ────
try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas is required.")
    print("  pip install pandas")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.")
    print("  pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

ROOT             = r"C:\dev\chaos"
DATA_DIR         = os.path.join(ROOT, "data")
MEAS_DIR         = os.path.join(ROOT, "measurements")
EXPERIMENTS_FILE = os.path.join(DATA_DIR, "experiments.json")
DEFAULT_OUTPUT   = os.path.join(DATA_DIR, "status_report.xlsx")

CONFIG_RE = re.compile(r"^th1_[pm]\d+_th2_[pm]\d+(_r\d+)?$")

LONG_RECORDING_CD = "th1_p180_th2_m179"

# The 8 canonical artefacts a fully-completed measurement should have.
CANONICAL_OUTPUTS = [
    "tracking.csv",
    "verification.csv",
    "debug.mp4",
    "phase_3d_trajectory.png",
    "phase_3d_rotation.mp4",
    "phase_animation.mp4",
    "phase_panels.png",
    "combined.mp4",
]

# Status fill colours.
FILL = {
    "DONE":         PatternFill("solid", fgColor="C6EFCE"),
    "TRACKED":      PatternFill("solid", fgColor="FFEB9C"),
    "IN PROGRESS":  PatternFill("solid", fgColor="FCE4D6"),
    "NEEDS WORK":   PatternFill("solid", fgColor="FFC7CE"),
    "DROP_HIGH":    PatternFill("solid", fgColor="FFC7CE"),
    "DROP_MED":     PatternFill("solid", fgColor="FFEB9C"),
    "HEADER":       PatternFill("solid", fgColor="D9E1F2"),
}

GREEN_FONT = Font(color="006100")
RED_FONT   = Font(color="9C0006")


# ─────────────────────────────────────────────
# REGISTRY HELPERS
# ─────────────────────────────────────────────

def synthesize_config_description(entry):
    """
    Build a config_description from theta1/theta2 (measured > release).
    Returns None if no angles available.
    """
    th1 = entry.get("theta1_measured")
    th2 = entry.get("theta2_measured")
    if th1 is None or th2 is None:
        th1 = entry.get("theta1_release")
        th2 = entry.get("theta2_release")
    if th1 is None or th2 is None:
        return None

    def fmt(angle):
        n = int(round(float(angle)))
        sign = "p" if n >= 0 else "m"
        return f"{sign}{abs(n):03d}"
    return f"th1_{fmt(th1)}_th2_{fmt(th2)}"


def resolve_config_description(key, entry):
    """Same precedence as the migration script."""
    cd = entry.get("config_description")
    if cd:
        return cd
    video = entry.get("video_file") or ""
    stem  = os.path.splitext(os.path.basename(video))[0]
    if CONFIG_RE.match(stem):
        return stem
    return synthesize_config_description(entry)


# ─────────────────────────────────────────────
# CSV-DERIVED METRICS
# ─────────────────────────────────────────────

def compute_dropout_metrics(csv_path):
    """
    Returns a dict: dropout_pct, green_dropout_pct, red_dropout_pct,
    free_frames, dropout_pattern, verification_suspect_frames (when a
    verification.csv exists alongside).
    Numerical rates are rounded to 2 decimals; pattern is a human string.
    """
    metrics = {
        "dropout_pct":       None,
        "green_dropout_pct": None,
        "red_dropout_pct":   None,
        "free_frames":       None,
        "dropout_pattern":   "—",
    }
    if not os.path.exists(csv_path):
        return metrics

    df = pd.read_csv(csv_path)
    free = df[df["phase"] == "free_swing"].copy().reset_index(drop=True)
    n = len(free)
    metrics["free_frames"] = int(n)
    if n == 0:
        metrics["dropout_pattern"] = "no free_swing frames"
        return metrics

    drops = free["dropout"].astype(int)
    metrics["dropout_pct"]       = round(100.0 * drops.sum() / n, 2)
    metrics["green_dropout_pct"] = round(
        100.0 * free["x_green"].isna().sum() / n, 2)
    metrics["red_dropout_pct"]   = round(
        100.0 * free["x_red"].isna().sum() / n, 2)

    metrics["dropout_pattern"] = describe_dropout_runs(free, drops)
    return metrics


def describe_dropout_runs(free, drops):
    """
    Build a human-readable string describing every contiguous dropout
    run in the free_swing phase. Examples:
      "none"
      "12 isolated frames"
      "3 runs: f201–f238 (3.35s–3.98s, 0.6s) | f890 (14.8s, isolated) | ..."
    Truncates to top 5 runs by duration when more than 10 exist.
    """
    runs = []   # list of (start_idx, end_idx, n)
    in_run = False
    s = 0
    for i, v in enumerate(drops.values):
        if v == 1 and not in_run:
            in_run, s = True, i
        elif v != 1 and in_run:
            runs.append((s, i - 1, i - s))
            in_run = False
    if in_run:
        runs.append((s, len(drops) - 1, len(drops) - s))

    if not runs:
        return "none"

    if all(n == 1 for _, _, n in runs):
        return f"{len(runs)} isolated frame{'s' if len(runs) != 1 else ''}"

    formatted = []
    for s_idx, e_idx, n in runs:
        f_start = int(free.loc[s_idx, "frame"])
        f_end   = int(free.loc[e_idx, "frame"])
        t_start = float(free.loc[s_idx, "time_s"])
        t_end   = float(free.loc[e_idx, "time_s"])
        dur = t_end - t_start
        if n == 1:
            formatted.append((dur,
                              f"f{f_start} ({t_start:.1f}s, isolated)"))
        else:
            formatted.append((dur,
                              f"f{f_start}–f{f_end} "
                              f"({t_start:.1f}s–{t_end:.1f}s, "
                              f"{dur:.1f}s)"))

    if len(runs) > 10:
        formatted.sort(key=lambda x: -x[0])
        head = formatted[:5]
        tail_n = len(runs) - 5
        body = " | ".join(s for _, s in head)
        return f"{len(runs)} runs (top 5 by duration): {body} ... and {tail_n} more"

    body = " | ".join(s for _, s in formatted)
    return f"{len(runs)} run{'s' if len(runs) != 1 else ''}: {body}"


def count_verification_suspects(verification_csv):
    """Count rows with suspect=1; returns None if file missing."""
    if not os.path.exists(verification_csv):
        return None
    try:
        df = pd.read_csv(verification_csv)
        if "suspect" not in df.columns:
            return None
        return int(df["suspect"].astype(int).sum())
    except Exception:
        return None


# ─────────────────────────────────────────────
# STATUS DERIVATION
# ─────────────────────────────────────────────

def derive_status(meas_dir, dropout_pct, tracking_quality):
    """
    Returns (status_str, status_bucket). status_bucket is one of the
    keys into FILL: "DONE", "TRACKED", "IN PROGRESS", "NEEDS WORK", or
    None for PENDING. The exact status_str is what the cell shows.
    """
    if not os.path.isdir(meas_dir):
        return "PENDING", None

    present = {name for name in CANONICAL_OUTPUTS
               if os.path.exists(os.path.join(meas_dir, name))}
    if not present:
        return "PENDING", None

    has_csv = "tracking.csv" in present

    if has_csv and dropout_pct is not None and dropout_pct > 10:
        return f"NEEDS WORK — high dropout ({dropout_pct:.1f}%)", "NEEDS WORK"

    if len(present) == len(CANONICAL_OUTPUTS):
        return "DONE ✅", "DONE"

    n_analysis = len(present - {"tracking.csv", "verification.csv", "debug.mp4"})
    if has_csv and n_analysis == 0:
        return "TRACKED — analysis pending", "TRACKED"

    if has_csv and tracking_quality != "verified":
        if n_analysis < 5:
            return (f"IN PROGRESS ({len(present)}/{len(CANONICAL_OUTPUTS)})",
                    "IN PROGRESS")
        return "TRACKED — verify pending", "TRACKED"

    return (f"IN PROGRESS ({len(present)}/{len(CANONICAL_OUTPUTS)})",
            "IN PROGRESS")


# ─────────────────────────────────────────────
# ROW BUILDER
# ─────────────────────────────────────────────

def build_row(key, entry, cd):
    """
    Build a dict of column-name → value for one measurement. The same
    keys are used by both sheets; the long_recording sheet appends
    extras from the caller.
    """
    meas_dir = os.path.join(MEAS_DIR, cd)
    csv_path = os.path.join(meas_dir, "tracking.csv")
    metrics  = compute_dropout_metrics(csv_path)

    # Tracked → use *_release; not yet tracked → fall back to *_measured.
    th1 = entry.get("theta1_release")
    th2 = entry.get("theta2_release")
    if th1 is None: th1 = entry.get("theta1_measured")
    if th2 is None: th2 = entry.get("theta2_measured")

    status_str, _ = derive_status(meas_dir,
                                  metrics["dropout_pct"],
                                  entry.get("tracking_quality"))

    presence = {name: ("✅" if os.path.exists(
                            os.path.join(meas_dir, name))
                       else "❌")
                for name in CANONICAL_OUTPUTS}

    # HSV calibration: per-video file overrides global.
    video_stem = os.path.splitext(os.path.basename(
        entry.get("video_file") or ""))[0]
    per_video_hsv = os.path.join(DATA_DIR, f"hsv_{video_stem}.json")
    hsv_label = "per-video" if (video_stem and
                                 os.path.exists(per_video_hsv)) else "global"

    notes_parts = []
    user_notes = (entry.get("notes") or "").strip()
    if user_notes:
        notes_parts.append(user_notes)

    combined_path = os.path.join(meas_dir, "combined.mp4")
    if os.path.exists(combined_path):
        if os.path.getsize(combined_path) < 1_000_000:
            notes_parts.append("⚠ combined.mp4 is suspiciously small (<1 MB)")

    if metrics["dropout_pct"] is not None and metrics["dropout_pct"] > 10:
        notes_parts.append("⚠ dropout > 10%")

    if entry.get("repeat_number") and int(entry["repeat_number"]) > 1:
        notes_parts.append("⚠ repeat trial")

    if (hsv_label == "global" and
            metrics["dropout_pct"] is not None and
            metrics["dropout_pct"] > 5):
        notes_parts.append("⚠ no per-video HSV calibration")

    if entry.get("verification_notes"):
        notes_parts.append("verification: " + entry["verification_notes"])

    return {
        "config_description": cd,
        "registry_key":       key,
        "video_file":         entry.get("video_file") or "",
        "repeat":             entry.get("repeat_number", 1) or 1,
        "theta1_ic":          th1 if th1 is not None else "",
        "theta2_ic":          th2 if th2 is not None else "",
        "status":             status_str,
        "dropout_%":          (metrics["dropout_pct"]
                                if metrics["dropout_pct"] is not None
                                else "—"),
        "green_dropout_%":    (metrics["green_dropout_pct"]
                                if metrics["green_dropout_pct"] is not None
                                else "—"),
        "red_dropout_%":      (metrics["red_dropout_pct"]
                                if metrics["red_dropout_pct"] is not None
                                else "—"),
        "free_frames":        (metrics["free_frames"]
                                if metrics["free_frames"] is not None
                                else "—"),
        "dropout_pattern":    metrics["dropout_pattern"],
        "tracking.csv":           presence["tracking.csv"],
        "verification.csv":       presence["verification.csv"],
        "debug.mp4":              presence["debug.mp4"],
        "phase_3d_trajectory":    presence["phase_3d_trajectory.png"],
        "phase_3d_rotation":      presence["phase_3d_rotation.mp4"],
        "phase_animation":        presence["phase_animation.mp4"],
        "phase_panels":           presence["phase_panels.png"],
        "combined.mp4":           presence["combined.mp4"],
        "hsv_config":         hsv_label,
        "notes":              "\n".join(notes_parts) if notes_parts else "",
    }


def add_long_recording_extras(row, entry, cd):
    """Tack on the four extra columns for the long_recording sheet."""
    meas_dir   = os.path.join(MEAS_DIR, cd)
    verif_csv  = os.path.join(meas_dir, "verification.csv")
    suspects   = count_verification_suspects(verif_csv)

    row["duration_s"]                  = entry.get("duration_s")
    row["n_free_frames"]                = entry.get("n_free_frames")
    row["verification_suspect_frames"] = (
        suspects if suspects is not None else "—")
    row["energy_proxy"]                = entry.get("energy_proxy")
    return row


# ─────────────────────────────────────────────
# WORKBOOK FORMATTING
# ─────────────────────────────────────────────

def write_sheet(ws, columns, rows, *, status_col="status",
                dropout_col="dropout_%",
                presence_cols=()):
    """
    Write a sheet with header + rows, then apply formatting:
      * bold blue header with wrap + 30px row height
      * column widths sized to content (12..45 chars)
      * status fill colour
      * dropout colour scale
      * green/red font on presence columns
      * frozen header + auto-filter
    """
    # Header
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = FILL["HEADER"]
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")
    ws.row_dimensions[1].height = 30

    # Data
    for row in rows:
        ws.append([row.get(c, "") for c in columns])

    n_rows = len(rows) + 1   # incl. header

    # Column widths
    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(col_name)
        for r in range(2, n_rows + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            longest_line = max((len(line) for line in s.splitlines()),
                               default=len(s))
            if longest_line > max_len:
                max_len = longest_line
        width = max(12, min(45, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Wrap long-text cells (notes, dropout_pattern).
    for col_name in ("notes", "dropout_pattern"):
        if col_name in columns:
            ci = columns.index(col_name) + 1
            for r in range(2, n_rows + 1):
                ws.cell(row=r, column=ci).alignment = Alignment(wrap_text=True,
                                                                 vertical="top")

    # Status fill
    if status_col in columns:
        ci = columns.index(status_col) + 1
        for r in range(2, n_rows + 1):
            cell = ws.cell(row=r, column=ci)
            v = str(cell.value or "")
            for bucket, fill in FILL.items():
                if bucket in ("HEADER", "DROP_HIGH", "DROP_MED"):
                    continue
                if v.startswith(bucket):
                    cell.fill = fill
                    break

    # Dropout colour scale
    if dropout_col in columns:
        ci = columns.index(dropout_col) + 1
        for r in range(2, n_rows + 1):
            cell = ws.cell(row=r, column=ci)
            v = cell.value
            if isinstance(v, (int, float)):
                if v > 10:
                    cell.fill = FILL["DROP_HIGH"]
                elif v >= 5:
                    cell.fill = FILL["DROP_MED"]

    # Presence columns: green tick / red cross
    for pc in presence_cols:
        if pc not in columns:
            continue
        ci = columns.index(pc) + 1
        for r in range(2, n_rows + 1):
            cell = ws.cell(row=r, column=ci)
            v = str(cell.value or "")
            if "✅" in v:
                cell.font = GREEN_FONT
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
            elif "❌" in v:
                cell.font = RED_FONT
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")

    # Freeze + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Build the measurement status report (xlsx).")
    p.add_argument("--output", default=DEFAULT_OUTPUT,
                   help=f"output path (default: {DEFAULT_OUTPUT})")
    return p.parse_args()


def main():
    args = parse_args()

    if not os.path.exists(EXPERIMENTS_FILE):
        print(f"ERROR: registry not found: {EXPERIMENTS_FILE}")
        return 2
    with open(EXPERIMENTS_FILE, "r", encoding="utf-8") as f:
        reg = json.load(f)

    if not os.path.isdir(MEAS_DIR):
        print(f"WARN: {MEAS_DIR} does not exist — every row will be PENDING.")

    main_rows = []
    long_row  = None
    for key in sorted(reg.keys()):
        entry = reg[key]
        cd = resolve_config_description(key, entry)
        if not cd:
            print(f"WARN: skipping '{key}' — could not resolve config_description")
            continue
        row = build_row(key, entry, cd)
        if cd == LONG_RECORDING_CD:
            long_row = add_long_recording_extras(row, entry, cd)
        else:
            main_rows.append(row)

    # Sort the main sheet by physics: theta1_ic asc, theta2_ic asc.
    def sort_key(r):
        t1 = r["theta1_ic"]
        t2 = r["theta2_ic"]
        return (
            float(t1) if isinstance(t1, (int, float)) else float("inf"),
            float(t2) if isinstance(t2, (int, float)) else float("inf"),
        )
    main_rows.sort(key=sort_key)

    main_columns = [
        "config_description", "registry_key", "video_file", "repeat",
        "theta1_ic", "theta2_ic", "status",
        "dropout_%", "green_dropout_%", "red_dropout_%",
        "free_frames", "dropout_pattern",
        "tracking.csv", "verification.csv", "debug.mp4",
        "phase_3d_trajectory", "phase_3d_rotation", "phase_animation",
        "phase_panels", "combined.mp4",
        "hsv_config", "notes",
    ]
    long_columns = main_columns + [
        "duration_s", "n_free_frames",
        "verification_suspect_frames", "energy_proxy",
    ]
    presence_cols = (
        "tracking.csv", "verification.csv", "debug.mp4",
        "phase_3d_trajectory", "phase_3d_rotation",
        "phase_animation", "phase_panels", "combined.mp4",
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Measurements"
    write_sheet(ws, main_columns, main_rows,
                presence_cols=presence_cols)

    long_ws = wb.create_sheet("long_recording")
    long_rows_list = [long_row] if long_row else []
    write_sheet(long_ws, long_columns, long_rows_list,
                presence_cols=presence_cols)

    out = args.output
    if not os.path.isabs(out):
        out = os.path.join(ROOT, out)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)

    # Summary line
    n_total   = len(main_rows) + (1 if long_row else 0)
    n_tracked = sum(1 for r in main_rows
                    if str(r["status"]).startswith(("TRACKED", "DONE",
                                                    "IN PROGRESS",
                                                    "NEEDS WORK")))
    if long_row and str(long_row["status"]).startswith(("TRACKED", "DONE",
                                                        "IN PROGRESS",
                                                        "NEEDS WORK")):
        n_tracked += 1
    n_done = sum(1 for r in main_rows + ([long_row] if long_row else [])
                 if str(r["status"]).startswith("DONE"))

    print(f"Report written: {os.path.relpath(out, ROOT)} — "
          f"{n_total} measurements, {n_tracked} tracked, {n_done} done")
    return 0


if __name__ == "__main__":
    sys.exit(main())
